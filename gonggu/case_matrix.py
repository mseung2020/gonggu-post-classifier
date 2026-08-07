#!/usr/bin/env python3
"""DB 전수 케이스 세분화 리포트 — 모든 상품 행을 여러 축으로 라벨링해서 O/X 표와
교차표, 그리고 "실제로 존재하는 모든 조합"의 카운트를 뽑는다.

읽기 전용이다(SELECT만 실행). DB에 아무것도 쓰지 않고 뷰도 만들지 않는다 —
`--emit-sql`로 같은 로직의 순수 SQL(뷰 DDL + 예제 GROUP BY)을 파일로 받을 수 있다.

실행:
    python3 -m gonggu.case_matrix                 # 전체 → data/output/case_matrix/*
    python3 -m gonggu.case_matrix --limit 200     # 소량 확인
    python3 -m gonggu.case_matrix --emit-sql      # 같은 로직을 queries/case_matrix.sql로

분석 단위는 **상품 행 1건**(gonggu_post_product / gonggu_video_product)이다.
link_status·상세수집(detail)이 상품 단위로 달리므로 부모(포스트/영상) 단위로 세면
"링크는 되고 상세는 안 된" 상태를 표현할 수 없다. 부모 단위 표는 별도 섹션으로 따로 낸다.

축(axis) 목록은 AXES에, O/X 플래그는 OX_FLAGS에 있다. 새 축을 넣고 싶으면
_axis_sql()에 CASE 하나 추가하고 AXES에 한 줄 넣으면 나머지 표는 자동으로 따라온다.
"""
import argparse
import collections
import csv
import pathlib
import re
import sys

import pymysql.cursors

from gonggu.common import ROOT, connect_dst

OUT_DIR = ROOT / 'data/output/case_matrix'

# ------------------------------------------------------------------
# 도메인 패턴 — candidate_url을 도메인 성격으로 분류하는 데만 쓴다.
# resolve_links/config.py의 BAD_DOMAINS/MALL_DOMAINS, linkbio_parser/hosts.py의
# 허브 목록, prompts.py의 url_type 정의에서 가져와 합쳤다(리포트 전용이라
# 파이프라인 판정에는 영향 없음 — 여기 목록을 늘려도 DB는 안 바뀐다).
# ------------------------------------------------------------------
P_WALL = ['nid.naver.com', 'accounts.kakao.com', 'account.kakao.com', 'mkt.shopping.naver']
P_HUB = ['litt.ly', 'lit.link', 'inpock', 'inpk.link', 'linktr.ee', '//tr.ee', 'hity.io',
         'instabio.cc', 'bio.site', 'linkon.id', 'linkseller.net', 'taplink', 'linkby.me']
P_BROKEN = ['/error', '/notfound', '/not-found', '/404']
P_FORM = ['docs.google.com/forms', 'forms.gle', 'form.naver.com', 'naver.me/form',
          'forms.office.com', 'tally.so', 'walla.my']
P_NSTORE = ['smartstore.naver.com', 'brand.naver.com']
P_NSHOP = ['shopping.naver.com']
P_NBLOG = ['blog.naver.com', 'cafe.naver.com']
P_SHORT = ['naver.me', 'bit.ly', 'srok.kr', 'me2.do', 'vo.la', 'buly.kr', 'link.coupang.com']
P_KAKAO = ['pf.kakao.com', 'open.kakao.com', 'kakao.com']
P_MARKET = ['coupang.com', 'gmarket.co.kr', 'auction.co.kr', '11st.co.kr', 'interpark.com',
            'tmon.co.kr', 'wemakeprice.com', 'lotteon.com', 'ssg.com', 'oliveyoung.co.kr',
            'kurly.com', 'musinsa.com', 'ably', 'zigzag']
P_PAY = ['srookpay', 'payapp', 'tosspayments', 'kakaopay', 'naverpay', 'smartro']
P_BUILDER = ['cafe24.com', 'imweb.me', 'shopby', 'sixshop', 'godomall', 'makeshop',
             'wixsite', 'myshopify.com']
# 상품 "상세" 페이지로 보이는 경로/파라미터 흔적. 스토어 메인·기획전·목록과 구분하는 용도.
P_ITEM = ['/products/', '/product/', '/goods', '/item', 'prod_no', 'productno', 'goodsno',
          '/vp/products/', '/detail', 'product_no', '/dp/']


def _any_like(col, pats):
    return '(' + ' OR '.join(f"{col} LIKE '%{p}%'" for p in pats) + ')'


def _no_url(col='f.candidate_url'):
    return f"({col} IS NULL OR {col} = '')"


# ------------------------------------------------------------------
# 1) 두 플랫폼을 같은 컬럼 이름으로 펼치는 SELECT (원본 컬럼명 차이를 여기서만 흡수)
# ------------------------------------------------------------------
def _flat_sql():
    ig = """
    SELECT 'ig' AS platform, pp.id AS product_id, p.post_id AS parent_key,
           p.user_id AS owner_id, p.publish_date AS publish_dt, NULL AS external_url,
           p.gonggu_start_date AS sd, p.gonggu_end_date AS ed, p.gonggu_stage AS stage,
           p.classification_note AS cnote,
           pp.product_name, pp.link_location, pp.url_type, pp.candidate_url,
           pp.link_status, pp.sort_order,
           (SELECT COUNT(*) FROM gonggu_post_product s WHERE s.post_id = p.post_id) AS sib_n,
           d.detail_status, d.thumbnail_url, d.brand_name_kr, d.brand_name_en,
           d.category, d.subcategory, d.search_keywords,
           d.original_price, d.sale_price, d.discount_rate, d.discount_amount,
           d.free_shipping, d.shipping_fee, d.shipping_note,
           d.composition_info, d.gift_info, d.coupon_info,
           d.ai_summary, d.ai_summary_confidence, d.detail_error,
           (SELECT COUNT(*) FROM gonggu_post_product_image i
              WHERE i.post_product_id = pp.id) AS img_n,
           (SELECT COUNT(*) FROM gonggu_post_product_image i
              WHERE i.post_product_id = pp.id AND i.image_type = 'thumbnail') AS img_thumb_n
      FROM gonggu_post p
      JOIN gonggu_post_product pp ON pp.post_id = p.post_id
      LEFT JOIN gonggu_post_product_detail d ON d.post_product_id = pp.id
    """
    yt = """
    SELECT 'yt' AS platform, vp.id AS product_id, v.video_id AS parent_key,
           v.channel_id AS owner_id, CAST(v.publishDate AS DATETIME) AS publish_dt,
           v.external_url AS external_url,
           v.gonggu_start_date AS sd, v.gonggu_end_date AS ed, v.gonggu_stage AS stage,
           v.classification_note AS cnote,
           vp.product_name, vp.link_location, vp.url_type, vp.candidate_url,
           vp.link_status, vp.sort_order,
           (SELECT COUNT(*) FROM gonggu_video_product s WHERE s.video_id = v.video_id) AS sib_n,
           d.detail_status, d.thumbnail_url, d.brand_name_kr, d.brand_name_en,
           d.category, d.subcategory, d.search_keywords,
           d.original_price, d.sale_price, d.discount_rate, d.discount_amount,
           d.free_shipping, d.shipping_fee, d.shipping_note,
           d.composition_info, d.gift_info, d.coupon_info,
           d.ai_summary, d.ai_summary_confidence, d.detail_error,
           (SELECT COUNT(*) FROM gonggu_video_product_image i
              WHERE i.video_product_id = vp.id) AS img_n,
           (SELECT COUNT(*) FROM gonggu_video_product_image i
              WHERE i.video_product_id = vp.id AND i.image_type = 'thumbnail') AS img_thumb_n
      FROM gonggu_video v
      JOIN gonggu_video_product vp ON vp.video_id = v.video_id
      LEFT JOIN gonggu_video_product_detail d ON d.video_product_id = vp.id
    """
    return f'({ig} UNION ALL {yt})'


# ------------------------------------------------------------------
# 2) 축(라벨) + O/X 플래그를 붙이는 SELECT
# ------------------------------------------------------------------
def _axis_sql():
    url_kind = f"""CASE
      WHEN {_no_url()}                       THEN '00_URL없음'
      WHEN {_any_like('f.candidate_url', P_WALL)}    THEN '01_로그인월·차단도메인'
      WHEN {_any_like('f.candidate_url', P_HUB)}     THEN '02_링크모음허브(인포크·litt.ly류)'
      WHEN {_any_like('f.candidate_url', P_BROKEN)}  THEN '03_깨진경로(error·404)'
      WHEN {_any_like('f.candidate_url', P_FORM)}    THEN '04_폼·설문'
      WHEN {_any_like('f.candidate_url', P_NSTORE)}  THEN '05_네이버_스마트·브랜드스토어'
      WHEN {_any_like('f.candidate_url', P_NSHOP)}   THEN '06_네이버_쇼핑'
      WHEN {_any_like('f.candidate_url', P_NBLOG)}   THEN '07_네이버_블로그·카페(몰아님)'
      WHEN {_any_like('f.candidate_url', P_SHORT)}   THEN '08_단축링크'
      WHEN {_any_like('f.candidate_url', P_KAKAO)}   THEN '09_카카오채널'
      WHEN {_any_like('f.candidate_url', P_MARKET)}  THEN '10_오픈마켓·대형몰'
      WHEN {_any_like('f.candidate_url', P_PAY)}     THEN '11_결제플랫폼'
      WHEN {_any_like('f.candidate_url', P_BUILDER)} THEN '12_자사몰(빌더도메인)'
      WHEN f.candidate_url LIKE '%naver.com%'        THEN '13_네이버_기타'
      ELSE '14_기타·독립도메인'
    END"""

    depth = f"""CASE
      WHEN {_no_url()} THEN 'N/A_URL없음'
      WHEN {_any_like('LOWER(f.candidate_url)', P_ITEM)} THEN '상품상세_추정'
      ELSE '메인·목록·기획전_추정'
    END"""

    unres = f"""CASE
      WHEN f.link_status = 'done' THEN 'R_해결됨(done)'
      WHEN {_no_url()} AND f.link_location = '링크없음_불명'                     THEN 'U01_링크자체없음(불명)'
      WHEN {_no_url()} AND f.link_location = '댓글참여_DM'                       THEN 'U02_댓글·DM유도_URL없음'
      WHEN {_no_url()} AND f.link_location = '고정댓글_더보기'                   THEN 'U03_고정댓글안내_URL없음'
      WHEN {_no_url()} AND f.link_location = '설명_프로필안내'                   THEN 'U04_프로필안내인데_URL없음'
      WHEN {_no_url()}                                                          THEN 'U05_직접링크라는데_URL없음'
      WHEN {_any_like('f.candidate_url', P_WALL)}                               THEN 'U06_로그인월·차단에서멈춤'
      WHEN {_any_like('f.candidate_url', P_HUB)}                                THEN 'U07_허브까지만(제품링크미확정)'
      WHEN {_any_like('f.candidate_url', P_BROKEN)}                             THEN 'U08_깨진경로'
      WHEN {_any_like('f.candidate_url', P_FORM)}                               THEN 'U09_폼·설문뿐'
      WHEN {_any_like('f.candidate_url', P_NBLOG)}                              THEN 'U10_블로그·카페뿐'
      WHEN {_any_like('f.candidate_url', P_SHORT)}                              THEN 'U11_단축링크_미해석'
      WHEN {_any_like('f.candidate_url', P_KAKAO)}                              THEN 'U12_카카오채널뿐'
      WHEN {_any_like('LOWER(f.candidate_url)', P_ITEM)}                        THEN 'U13_상품상세로보이는데_미확정'
      ELSE 'U14_몰·기타도메인_메인·목록에서멈춤'
    END"""

    return f"""
    SELECT f.*,
      /* --- 축(라벨) --- */
      CASE WHEN f.sd IS NOT NULL AND f.ed IS NOT NULL THEN 'A_시작O_종료O'
           WHEN f.sd IS NOT NULL AND f.ed IS     NULL THEN 'B_시작O_종료X'
           WHEN f.sd IS     NULL AND f.ed IS NOT NULL THEN 'C_시작X_종료O'
           ELSE                                            'D_시작X_종료X' END AS c_period,
      COALESCE(f.stage, '(NULL)')                                 AS c_stage,
      COALESCE(f.link_location, '(NULL)')                         AS c_loc,
      COALESCE(NULLIF(f.url_type, ''), '(NULL)')                  AS c_urltype,
      {url_kind}                                                  AS c_urlkind,
      {depth}                                                     AS c_depth,
      COALESCE(f.link_status, '(NULL_미해석)')                     AS c_linkstatus,
      COALESCE(f.detail_status, '(행없음)')                        AS c_detail,
      {unres}                                                     AS c_unres,
      CASE WHEN f.sib_n > 1 THEN '다상품(2+)' ELSE '단일상품' END   AS c_multi,
      CASE WHEN f.platform = 'ig' THEN 'N/A_인스타'
           WHEN f.external_url IS NULL OR f.external_url = '' THEN '채널링크없음'
           ELSE '채널링크있음' END                                 AS c_ext,
      /* --- O/X 플래그 (1=O, 0=X) --- */
      CASE WHEN f.sd IS NOT NULL THEN 1 ELSE 0 END                                    AS ox_start,
      CASE WHEN f.ed IS NOT NULL THEN 1 ELSE 0 END                                    AS ox_end,
      CASE WHEN f.cnote IS NOT NULL AND f.cnote <> '' THEN 1 ELSE 0 END               AS ox_note,
      CASE WHEN {_no_url()} THEN 0 ELSE 1 END                                         AS ox_url,
      CASE WHEN f.link_status = 'done' THEN 1 ELSE 0 END                              AS ox_linkdone,
      CASE WHEN {_any_like('f.candidate_url', P_NSTORE + P_NSHOP)} THEN 1 ELSE 0 END  AS ox_naver,
      CASE WHEN {_any_like('f.candidate_url', P_HUB)} THEN 1 ELSE 0 END               AS ox_hub,
      CASE WHEN f.detail_status IS NOT NULL THEN 1 ELSE 0 END                         AS ox_detailrow,
      CASE WHEN f.detail_status = 'done' THEN 1 ELSE 0 END                            AS ox_detaildone,
      CASE WHEN f.sale_price IS NOT NULL THEN 1 ELSE 0 END                            AS ox_price,
      CASE WHEN f.original_price IS NOT NULL THEN 1 ELSE 0 END                        AS ox_origprice,
      CASE WHEN f.discount_rate IS NOT NULL OR f.discount_amount IS NOT NULL
                THEN 1 ELSE 0 END                                                     AS ox_discount,
      CASE WHEN f.brand_name_kr IS NOT NULL OR f.brand_name_en IS NOT NULL
                THEN 1 ELSE 0 END                                                     AS ox_brand,
      CASE WHEN f.category IS NOT NULL THEN 1 ELSE 0 END                              AS ox_category,
      CASE WHEN f.search_keywords IS NOT NULL THEN 1 ELSE 0 END                       AS ox_keywords,
      CASE WHEN f.free_shipping IS NOT NULL OR f.shipping_fee IS NOT NULL
                OR (f.shipping_note IS NOT NULL AND f.shipping_note <> '')
                THEN 1 ELSE 0 END                                                     AS ox_shipping,
      CASE WHEN f.composition_info IS NOT NULL OR f.gift_info IS NOT NULL
                OR f.coupon_info IS NOT NULL THEN 1 ELSE 0 END                        AS ox_promo,
      CASE WHEN f.thumbnail_url IS NOT NULL AND f.thumbnail_url <> '' THEN 1 ELSE 0 END AS ox_thumb,
      CASE WHEN f.img_n > 0 THEN 1 ELSE 0 END                                         AS ox_image,
      CASE WHEN f.ai_summary IS NOT NULL AND f.ai_summary <> '' THEN 1 ELSE 0 END      AS ox_summary,
      CASE WHEN f.ai_summary_confidence >= 70 THEN 1 ELSE 0 END                       AS ox_conf70
    FROM {_flat_sql()} f
    """


# 축: (컬럼, 사람이 읽는 이름)
AXES = [
    ('platform', '플랫폼'),
    ('c_period', '공구기간_NULL패턴'),
    ('c_stage', 'gonggu_stage'),
    ('c_loc', 'link_location'),
    ('c_urltype', 'url_type(LLM판정)'),
    ('c_urlkind', 'candidate_url_도메인유형'),
    ('c_depth', 'URL_상세페이지여부'),
    ('c_linkstatus', 'link_status'),
    ('c_detail', 'detail_status'),
    ('c_unres', '미해결_세부사유'),
    ('c_multi', '부모_상품수'),
    ('c_ext', '유튜브_채널링크'),
]

# O/X 플래그: (컬럼, 표 머리글)
OX_FLAGS = [
    ('ox_start', '시작일'),
    ('ox_end', '종료일'),
    ('ox_note', '특이사항'),
    ('ox_url', '후보URL'),
    ('ox_linkdone', '링크확정'),
    ('ox_naver', '네이버링크'),
    ('ox_hub', '허브링크'),
    ('ox_detailrow', '상세행'),
    ('ox_detaildone', '상세완료'),
    ('ox_price', '판매가'),
    ('ox_origprice', '정가'),
    ('ox_discount', '할인'),
    ('ox_brand', '브랜드'),
    ('ox_category', '카테고리'),
    ('ox_keywords', '키워드'),
    ('ox_shipping', '배송'),
    ('ox_promo', '구성·사은·쿠폰'),
    ('ox_thumb', '썸네일'),
    ('ox_image', '이미지'),
    ('ox_summary', 'AI요약'),
    ('ox_conf70', '신뢰도70+'),
]

# 리포트 정의서에 그대로 찍히는 "O의 뜻" — 표를 보는 사람이 O/X 기준을 되짚을 수 있게.
_OX_DEF = {
    'ox_start': 'gonggu_start_date NOT NULL',
    'ox_end': 'gonggu_end_date NOT NULL',
    'ox_note': 'classification_note 있음',
    'ox_url': 'candidate_url 있음',
    'ox_linkdone': "link_status='done'",
    'ox_naver': 'candidate_url이 네이버 스마트·브랜드스토어/쇼핑',
    'ox_hub': 'candidate_url이 링크모음 허브(인포크·litt.ly류)',
    'ox_detailrow': 'detail 테이블에 행이 있음',
    'ox_detaildone': "detail_status='done'",
    'ox_price': 'sale_price 있음',
    'ox_origprice': 'original_price 있음',
    'ox_discount': 'discount_rate 또는 discount_amount 있음',
    'ox_brand': 'brand_name_kr 또는 brand_name_en 있음',
    'ox_category': 'category 있음',
    'ox_keywords': 'search_keywords 있음',
    'ox_shipping': 'free_shipping/shipping_fee/shipping_note 중 하나 있음',
    'ox_promo': 'composition_info/gift_info/coupon_info 중 하나 있음',
    'ox_thumb': 'thumbnail_url 있음',
    'ox_image': '_image 테이블에 이미지 1장 이상',
    'ox_summary': 'ai_summary 있음',
    'ox_conf70': 'ai_summary_confidence >= 70',
}

# 교차표: (행축, 열축, 제목)
CROSSTABS = [
    ('c_linkstatus', 'c_urlkind', 'link_status × candidate_url 도메인유형'),
    ('c_linkstatus', 'c_period', 'link_status × 공구기간 NULL패턴'),
    ('c_linkstatus', 'c_loc', 'link_status × link_location'),
    ('c_linkstatus', 'c_depth', 'link_status × URL 상세페이지여부'),
    ('c_stage', 'c_period', 'gonggu_stage × 공구기간 NULL패턴  (종료+종료일X = 강제종료 추정)'),
    ('c_stage', 'c_linkstatus', 'gonggu_stage × link_status'),
    ('c_urltype', 'c_urlkind', 'url_type(LLM) × 실제 도메인유형  (불일치 = LLM 오분류 후보)'),
    ('c_detail', 'c_linkstatus', 'detail_status × link_status'),
    ('c_detail', 'c_urlkind', 'detail_status × 도메인유형'),
    ('c_unres', 'platform', '미해결 세부사유 × 플랫폼'),
    ('c_unres', 'c_stage', '미해결 세부사유 × gonggu_stage'),
    ('c_loc', 'c_urlkind', 'link_location × 도메인유형'),
]

# "전체 조합" 표에 쓰는 축 묶음
COMBO_CORE = ['platform', 'c_period', 'c_stage', 'c_loc', 'c_urlkind', 'c_linkstatus', 'c_detail']
COMBO_FULL = [c for c, _ in AXES]


# ------------------------------------------------------------------
# 표 렌더링 헬퍼
# ------------------------------------------------------------------
def _md_table(headers, rows):
    out = ['| ' + ' | '.join(str(h) for h in headers) + ' |',
           '|' + '|'.join('---' for _ in headers) + '|']
    for r in rows:
        out.append('| ' + ' | '.join('' if v is None else str(v) for v in r) + ' |')
    return '\n'.join(out)


def _pct(n, total):
    return f'{100.0 * n / total:.1f}%' if total else '-'


def _dist(rows, col):
    c = collections.Counter(r[col] for r in rows)
    return sorted(c.items(), key=lambda kv: (-kv[1], str(kv[0])))


def _crosstab(rows, rax, cax):
    cells = collections.Counter((r[rax], r[cax]) for r in rows)
    rkeys = sorted({k[0] for k in cells}, key=str)
    ckeys = sorted({k[1] for k in cells}, key=str)
    return rkeys, ckeys, cells


def _combo(rows, cols):
    c = collections.Counter(tuple(r[col] for col in cols) for r in rows)
    return sorted(c.items(), key=lambda kv: (-kv[1], kv[0]))


# ------------------------------------------------------------------
def build_report(rows):
    n = len(rows)
    parents = {(r['platform'], r['parent_key']) for r in rows}
    md = ['# 공구왕 DB 케이스 전수 세분화 리포트', '',
          f'- 분석 단위: **상품 행 1건** — 총 **{n:,}건**',
          f'- 부모(포스트/영상) 수: **{len(parents):,}건**',
          f"  - 인스타 포스트 {len({p for p in parents if p[0] == 'ig'}):,} / "
          f"유튜브 영상 {len({p for p in parents if p[0] == 'yt'}):,}",
          '- O/X 표기: **O = 값이 있음(NOT NULL·비어있지 않음)**, X = 없음',
          '', '---', '']

    # 0) 케이스 정의서 — 어떤 축으로 쪼갰는지 한눈에
    md += ['## 0. 세분화 축 정의서', '',
           _md_table(['#', '축', '컬럼', '실제 등장 값 수', '값 목록'],
                     [[i, name, f'`{col}`', len(_dist(rows, col)),
                       ' / '.join(str(k) for k, _ in _dist(rows, col))]
                      for i, (col, name) in enumerate(AXES, 1)]), '',
           _md_table(['#', 'O/X 플래그', '컬럼', 'O의 정의'],
                     [[i, h, f'`{c}`', _OX_DEF.get(c, '')]
                      for i, (c, h) in enumerate(OX_FLAGS, 1)]), '', '---', '']

    # 1) 축별 단독 분포
    md += ['## 1. 축별 단독 분포', '']
    for i, (col, name) in enumerate(AXES, 1):
        d = _dist(rows, col)
        md += [f'### 1.{i} {name}  (`{col}`)', '',
               _md_table(['값', '건수', '비율'],
                         [[k, f'{v:,}', _pct(v, n)] for k, v in d]
                         + [['**합계**', f'**{n:,}**', '100%']]), '']

    # 2) 교차표
    md += ['---', '', '## 2. 2축 교차표', '']
    for i, (rax, cax, title) in enumerate(CROSSTABS, 1):
        rkeys, ckeys, cells = _crosstab(rows, rax, cax)
        headers = [f'{dict(AXES)[rax]} ↓ / {dict(AXES)[cax]} →'] + ckeys + ['합계']
        body = []
        for rk in rkeys:
            line = [rk] + [f'{cells[(rk, ck)]:,}' if cells[(rk, ck)] else '·' for ck in ckeys]
            line.append(f'**{sum(cells[(rk, ck)] for ck in ckeys):,}**')
            body.append(line)
        tot = ['**합계**'] + [f'**{sum(cells[(rk, ck)] for rk in rkeys):,}**' for ck in ckeys] \
              + [f'**{n:,}**']
        body.append(tot)
        md += [f'### 2.{i} {title}', '', _md_table(headers, body), '']

    # 3) O/X 조합 표 (핵심 요청)
    md += ['---', '', '## 3. O/X 조합 표 — 실제 존재하는 조합만',
           '', '한 줄이 하나의 "케이스"다. 21개 플래그 조합이 전부 다르면 전부 다른 줄로 나온다.', '']
    ox_cols = [c for c, _ in OX_FLAGS]
    ox_combo = _combo(rows, ox_cols)
    md += [f'- 이론상 조합 수 2^{len(ox_cols)} 중 **실제 등장 조합 {len(ox_combo):,}가지**', '']
    headers = [h for _, h in OX_FLAGS] + ['건수', '비율']
    body = [[('O' if v else 'X') for v in key] + [f'{cnt:,}', _pct(cnt, n)]
            for key, cnt in ox_combo[:60]]
    md += [_md_table(headers, body), '']
    if len(ox_combo) > 60:
        md += [f'> 상위 60개만 표시 — 전체 {len(ox_combo):,}가지는 `case_matrix_ox.csv` 참고', '']

    # 4) 전체 라벨 조합
    md += ['---', '', '## 4. 라벨 조합 표 — 실제 존재하는 조합만', '']
    for tag, cols, title in [('4.1', COMBO_CORE, '핵심 7축 조합'),
                             ('4.2', COMBO_FULL, f'최대 세분화 {len(COMBO_FULL)}축 조합')]:
        combo = _combo(rows, cols)
        md += [f'### {tag} {title}', '',
               f'- 축: {", ".join(dict(AXES)[c] for c in cols)}',
               f'- **실제 등장 조합 {len(combo):,}가지**', '',
               _md_table([dict(AXES)[c] for c in cols] + ['건수', '비율'],
                         [list(k) + [f'{v:,}', _pct(v, n)] for k, v in combo[:80]]), '']
        if len(combo) > 80:
            md += [f'> 상위 80개만 표시 — 전체는 CSV 참고', '']

    # 5) 부모 단위 표
    md += ['---', '', '## 5. 부모(포스트/영상) 단위 표', '',
           '상품 행이 아니라 부모 1건 기준. 같은 부모가 여러 상품을 가지면 1건으로 센다.', '']
    pmap = {}
    for r in rows:
        pmap.setdefault((r['platform'], r['parent_key']), r)
    prows = list(pmap.values())
    for i, col in enumerate(['platform', 'c_period', 'c_stage', 'c_multi'], 1):
        d = _dist(prows, col)
        md += [f'### 5.{i} {dict(AXES)[col]} (부모 단위)', '',
               _md_table(['값', '건수', '비율'],
                         [[k, f'{v:,}', _pct(v, len(prows))] for k, v in d]), '']
    rkeys, ckeys, cells = _crosstab(prows, 'c_stage', 'c_period')
    md += ['### 5.5 gonggu_stage × 공구기간 (부모 단위)', '',
           _md_table(['stage ↓ / 기간 →'] + ckeys,
                     [[rk] + [f'{cells[(rk, ck)]:,}' if cells[(rk, ck)] else '·'
                              for ck in ckeys] for rk in rkeys]), '']
    return '\n'.join(md)


def write_csvs(rows, out_dir=None):
    OUT_DIR = out_dir or globals()['OUT_DIR']          # noqa: F841 (아래 경로 조립에 씀)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    n = len(rows)

    # (a) 상품 1건 = 1행 원본 라벨링 결과
    axis_cols = [c for c, _ in AXES]
    ox_cols = [c for c, _ in OX_FLAGS]
    base = ['platform', 'parent_key', 'product_id', 'owner_id', 'product_name',
            'candidate_url', 'link_status', 'detail_status', 'sd', 'ed', 'stage']
    with (OUT_DIR / 'case_matrix_rows.csv').open('w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        w.writerow(base + axis_cols + ox_cols)
        for r in rows:
            w.writerow([r[c] for c in base] + [r[c] for c in axis_cols]
                       + [('O' if r[c] else 'X') for c in ox_cols])

    # (b) O/X 조합 전수
    with (OUT_DIR / 'case_matrix_ox.csv').open('w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        w.writerow([h for _, h in OX_FLAGS] + ['건수', '비율'])
        for key, cnt in _combo(rows, ox_cols):
            w.writerow([('O' if v else 'X') for v in key] + [cnt, _pct(cnt, n)])

    # (c) 라벨 조합 전수(최대 세분화)
    with (OUT_DIR / 'case_matrix_combo.csv').open('w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        w.writerow([dict(AXES)[c] for c in COMBO_FULL] + ['건수', '비율'])
        for key, cnt in _combo(rows, COMBO_FULL):
            w.writerow(list(key) + [cnt, _pct(cnt, n)])

    # (d) 축별 단독 분포 한 파일에
    with (OUT_DIR / 'case_matrix_dist.csv').open('w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        w.writerow(['축', '값', '건수', '비율'])
        for col, name in AXES:
            for k, v in _dist(rows, col):
                w.writerow([name, k, v, _pct(v, n)])


def load_rows_from_csv(path):
    """앞선 실행이 남긴 case_matrix_rows.csv에서 rows를 복원한다 — DB에 다시 붙지 않고
    엑셀/리포트만 다시 만들 때 쓴다. O/X 문자는 다시 1/0으로 되돌려서 집계 코드가
    DB에서 직접 읽은 경우와 완전히 같은 자료구조를 보게 한다."""
    ox = {c for c, _ in OX_FLAGS}
    rows = []
    with pathlib.Path(path).open(newline='', encoding='utf-8-sig') as fh:
        for r in csv.DictReader(fh):
            rows.append({k: (1 if v == 'O' else 0) if k in ox else v for k, v in r.items()})
    return rows


# ------------------------------------------------------------------
# 엑셀 — 시트 하나 = 표 하나. 축을 늘리면 시트도 자동으로 따라 늘어난다.
# ------------------------------------------------------------------
_XL_HDR_FILL = 'FF2F4858'
_XL_SUB_FILL = 'FFE8EEF2'

# 엑셀 시트명은 31자 제한이라 교차표 시트에는 축 별명을 쓴다(시트 안 1행에 정식 제목을 남김).
_AXIS_SHORT = {
    'platform': '플랫폼', 'c_period': '기간', 'c_stage': 'stage', 'c_loc': '링크위치',
    'c_urltype': 'urltype', 'c_urlkind': '도메인', 'c_depth': '상세페이지',
    'c_linkstatus': '링크상태', 'c_detail': '상세상태', 'c_unres': '미해결사유',
    'c_multi': '상품수', 'c_ext': '채널링크',
}


def _xl_style(ws, header_row=1, freeze='A2', widths=None, autofilter=True):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    fill = PatternFill('solid', fgColor=_XL_HDR_FILL)
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = freeze
    if autofilter and ws.max_row > header_row:
        ws.auto_filter.ref = (f'A{header_row}:'
                             f'{get_column_letter(ws.max_column)}{ws.max_row}')
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[letter].width = widths[i - 1]
            continue
        longest = max((len(str(ws.cell(row=r, column=i).value or ''))
                       for r in range(header_row, min(ws.max_row, header_row + 400) + 1)),
                      default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), 46)


def _xl_sheet(wb, title):
    return wb.create_sheet(re.sub(r'[\[\]:*?/\\]', '·', title)[:31])


def _xl_num(ws, cols, header_row=1):
    for c in cols:
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = '#,##0'


def write_xlsx(rows, path):
    """모든 표를 시트로 쪼갠 워크북 하나. 표만 담고 서식은 최소(머리글·고정창·필터·천단위)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    n = len(rows)
    parents = {(r['platform'], r['parent_key']) for r in rows}
    wb = Workbook()
    wb.remove(wb.active)

    # --- 00 요약 + 축 정의서 ---
    ws = _xl_sheet(wb, '00_요약·축정의서')
    ws.append(['공구왕 DB 케이스 전수 세분화'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    for label, val in [('상품 행(분석 단위)', n), ('부모(포스트/영상)', len(parents)),
                       ('인스타 포스트', len({p for p in parents if p[0] == 'ig'})),
                       ('유튜브 영상', len({p for p in parents if p[0] == 'yt'})),
                       ('라벨 축 수', len(AXES)), ('O/X 플래그 수', len(OX_FLAGS)),
                       ('O/X 조합 실존', len(_combo(rows, [c for c, _ in OX_FLAGS]))),
                       ('최대세분화 조합 실존', len(_combo(rows, COMBO_FULL)))]:
        ws.append([label, val])
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).number_format = '#,##0'
    ws.append([])
    hdr = ws.max_row + 1
    ws.append(['#', '라벨 축', '컬럼', '값 수', '값 목록'])
    for i, (col, name) in enumerate(AXES, 1):
        d = _dist(rows, col)
        ws.append([i, name, col, len(d), ' / '.join(str(k) for k, _ in d)])
    for cell in ws[hdr]:
        cell.font, cell.fill = Font(bold=True, color='FFFFFFFF'), \
            PatternFill('solid', fgColor=_XL_HDR_FILL)
    ws.append([])
    hdr2 = ws.max_row + 1
    ws.append(['#', 'O/X 플래그', '컬럼', 'O의 정의'])
    for i, (col, h) in enumerate(OX_FLAGS, 1):
        ws.append([i, h, col, _OX_DEF.get(col, '')])
    for cell in ws[hdr2]:
        cell.font, cell.fill = Font(bold=True, color='FFFFFFFF'), \
            PatternFill('solid', fgColor=_XL_HDR_FILL)
    for i, w in enumerate([6, 26, 20, 8, 90], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A3'
    summary = ws

    # --- 01 축별 단독 분포 ---
    ws = _xl_sheet(wb, '01_축별분포')
    ws.append(['라벨 축', '값', '건수', '비율(%)'])
    for col, name in AXES:
        for k, v in _dist(rows, col):
            ws.append([name, k, v, round(100.0 * v / n, 2)])
    _xl_style(ws, widths=[26, 46, 12, 10])
    _xl_num(ws, [3])

    # --- 02~ 교차표(축 2개짜리) ---
    for i, (rax, cax, title) in enumerate(CROSSTABS, 1):
        rkeys, ckeys, cells = _crosstab(rows, rax, cax)
        ws = _xl_sheet(wb, f'X{i:02d}_{_AXIS_SHORT[rax]}×{_AXIS_SHORT[cax]}')
        ws.append([title])
        ws['A1'].font = Font(bold=True)
        ws.append([f'{dict(AXES)[rax]} ↓ / {dict(AXES)[cax]} →'] + ckeys + ['합계'])
        for rk in rkeys:
            vals = [cells[(rk, ck)] for ck in ckeys]
            ws.append([rk] + [v or None for v in vals] + [sum(vals)])
        ws.append(['합계'] + [sum(cells[(rk, ck)] for rk in rkeys) or None for ck in ckeys] + [n])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for r in range(3, ws.max_row + 1):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
        _xl_style(ws, header_row=2, freeze='B3',
                  widths=[38] + [15] * (len(ckeys) + 1), autofilter=False)
        _xl_num(ws, range(2, ws.max_column + 1), header_row=2)

    # --- 30 O/X 조합 전수 ---
    ws = _xl_sheet(wb, '30_OX조합_전수')
    ws.append([h for _, h in OX_FLAGS] + ['건수', '비율(%)'])
    ox_cols = [c for c, _ in OX_FLAGS]
    o_fill = PatternFill('solid', fgColor='FFD8EFD3')
    for key, cnt in _combo(rows, ox_cols):
        ws.append([('O' if v else 'X') for v in key] + [cnt, round(100.0 * cnt / n, 2)])
        for j, v in enumerate(key, 1):
            c = ws.cell(row=ws.max_row, column=j)
            c.alignment = Alignment(horizontal='center')
            if v:
                c.fill = o_fill
    _xl_style(ws, widths=[7] * len(ox_cols) + [12, 10])
    _xl_num(ws, [len(ox_cols) + 1])

    # --- 20/31 라벨 조합 ---
    for name, cols in [('20_핵심조합', COMBO_CORE), ('31_전체조합_최대세분화', COMBO_FULL)]:
        ws = _xl_sheet(wb, name)
        ws.append([dict(AXES)[c] for c in cols] + ['건수', '비율(%)', '누적비율(%)'])
        run = 0
        for key, cnt in _combo(rows, cols):
            run += cnt
            ws.append(list(key) + [cnt, round(100.0 * cnt / n, 2), round(100.0 * run / n, 2)])
        _xl_style(ws, widths=[24] * len(cols) + [12, 10, 12])
        _xl_num(ws, [len(cols) + 1])

    # --- 40 부모 단위 ---
    pmap = {}
    for r in rows:
        pmap.setdefault((r['platform'], r['parent_key']), r)
    prows = list(pmap.values())
    ws = _xl_sheet(wb, '40_부모단위')
    ws.append(['라벨 축', '값', '건수', '비율(%)'])
    for col in ['platform', 'c_period', 'c_stage', 'c_multi']:
        for k, v in _dist(prows, col):
            ws.append([dict(AXES)[col], k, v, round(100.0 * v / len(prows), 2)])
    _xl_style(ws, widths=[26, 40, 12, 10])
    _xl_num(ws, [3])

    # --- 50 미해결 진단 (done 아닌 행만, 사유×상태×기간) ---
    bad = [r for r in rows if not r['ox_linkdone']]
    ws = _xl_sheet(wb, '50_미해결진단')
    ws.append(['미해결 세부사유', 'link_status', '공구기간', '플랫폼', 'link_location',
               '건수', '미해결중비율(%)'])
    if bad:
        for key, cnt in _combo(bad, ['c_unres', 'c_linkstatus', 'c_period', 'platform', 'c_loc']):
            ws.append(list(key) + [cnt, round(100.0 * cnt / len(bad), 2)])
    _xl_style(ws, widths=[34, 16, 18, 10, 20, 12, 14])
    _xl_num(ws, [6])

    # --- 60 상세수집 진행 (link_status=done 기준) ---
    okl = [r for r in rows if r['ox_linkdone']]
    ws = _xl_sheet(wb, '60_상세수집진행')
    ws.append(['도메인유형', 'URL_상세페이지여부', 'detail_status', '건수',
               'done중비율(%)'])
    if okl:
        for key, cnt in _combo(okl, ['c_urlkind', 'c_depth', 'c_detail']):
            ws.append(list(key) + [cnt, round(100.0 * cnt / len(okl), 2)])
    _xl_style(ws, widths=[34, 24, 16, 12, 14])
    _xl_num(ws, [4])

    # --- 00 시트 맨 아래에 목차 (시트가 20개라 어디에 뭐가 있는지 먼저 보이게) ---
    xdesc = {f'X{i:02d}': t for i, (_, _, t) in enumerate(CROSSTABS, 1)}
    pdesc = {'01_': '축별 단독 분포 — 모든 라벨 축의 값·건수·비율',
             '20_': f'핵심 {len(COMBO_CORE)}축 조합 전수',
             '30_': f'O/X {len(OX_FLAGS)}플래그 조합 전수 (한 줄 = 한 케이스)',
             '31_': f'최대 세분화 {len(COMBO_FULL)}축 조합 전수 + 누적비율',
             '40_': '부모(포스트/영상) 단위 분포',
             '50_': '미해결(done 아님)만 사유×상태×기간×플랫폼×링크위치',
             '60_': '링크확정(done) 건의 상세수집 진행 상황'}
    r0 = summary.max_row + 2
    summary.cell(row=r0, column=1, value='시트 목차').font = Font(bold=True, size=12)
    for j, h in enumerate(['시트', '표 행수', '내용'], 1):
        c = summary.cell(row=r0 + 1, column=j, value=h)
        c.font, c.fill = Font(bold=True, color='FFFFFFFF'), \
            PatternFill('solid', fgColor=_XL_HDR_FILL)
    for k, name in enumerate(wb.sheetnames):
        d = xdesc.get(name[:3]) or pdesc.get(name[:3]) or '요약·축 정의서'
        summary.cell(row=r0 + 2 + k, column=1, value=name)
        summary.cell(row=r0 + 2 + k, column=2, value=wb[name].max_row)
        summary.cell(row=r0 + 2 + k, column=3, value=d)

    pathlib.Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _one_sheet_ops(rows):
    """단일 시트에 세로로 쌓을 블록들을 순서대로 만든다 — 실제 쓰기는 write_xlsx_one이 한다.
    ('h2', 제목) / ('note', 설명) / ('tbl', 머리글, 본문, 옵션) / ('gap',) 의 리스트.
    쓰기와 분리해 둔 이유는 목차의 행 번호를 미리 계산해야 하기 때문(1패스로는 못 씀)."""
    n = len(rows)
    parents = {(r['platform'], r['parent_key']) for r in rows}
    ops = []

    def h2(t):
        ops.append(('h2', t))

    ops.append(('h2', '§0. 요약'))
    ops.append(('tbl', ['항목', '값'],
                [['상품 행(분석 단위)', n], ['부모(포스트/영상)', len(parents)],
                 ['  인스타 포스트', len({p for p in parents if p[0] == 'ig'})],
                 ['  유튜브 영상', len({p for p in parents if p[0] == 'yt'})],
                 ['라벨 축 수', len(AXES)], ['O/X 플래그 수', len(OX_FLAGS)],
                 ['O/X 조합 실존', len(_combo(rows, [c for c, _ in OX_FLAGS]))],
                 ['최대세분화 조합 실존', len(_combo(rows, COMBO_FULL))]],
                {'num': [2], 'boldfirstcol': True}))
    ops.append(('gap',))

    h2('§1. 라벨 축 정의서')
    ops.append(('tbl', ['#', '라벨 축', '컬럼', '값 수', '값 목록'],
                [[i, name, col, len(_dist(rows, col)),
                  ' / '.join(str(k) for k, _ in _dist(rows, col))]
                 for i, (col, name) in enumerate(AXES, 1)], {'num': [4]}))
    ops.append(('gap',))

    h2('§2. O/X 플래그 정의  (O = 값이 있음)')
    ops.append(('tbl', ['#', 'O/X 플래그', '컬럼', 'O의 정의'],
                [[i, h, col, _OX_DEF.get(col, '')]
                 for i, (col, h) in enumerate(OX_FLAGS, 1)], {}))
    ops.append(('gap',))

    h2('§3. 축별 단독 분포')
    for i, (col, name) in enumerate(AXES, 1):
        ops.append(('note', f'3.{i} {name}  ({col})'))
        d = _dist(rows, col)
        ops.append(('tbl', ['값', '건수', '비율(%)'],
                    [[k, v, round(100.0 * v / n, 2)] for k, v in d]
                    + [['합계', n, 100.0]],
                    {'num': [2], 'boldlast': True}))
        ops.append(('gap',))

    h2('§4. 2축 교차표')
    for i, (rax, cax, title) in enumerate(CROSSTABS, 1):
        rkeys, ckeys, cells = _crosstab(rows, rax, cax)
        ops.append(('note', f'4.{i} {title}'))
        body = []
        for rk in rkeys:
            vals = [cells[(rk, ck)] for ck in ckeys]
            body.append([rk] + [v or None for v in vals] + [sum(vals)])
        body.append(['합계'] + [sum(cells[(rk, ck)] for rk in rkeys) or None
                              for ck in ckeys] + [n])
        ops.append(('tbl', [f'{_AXIS_SHORT[rax]} ↓ / {_AXIS_SHORT[cax]} →'] + ckeys + ['합계'],
                    body, {'num': list(range(2, len(ckeys) + 3)),
                           'boldlast': True, 'boldfirstcol': True}))
        ops.append(('gap',))

    ox_cols = [c for c, _ in OX_FLAGS]
    ox_combo = _combo(rows, ox_cols)
    h2(f'§5. O/X 조합 전수 — 한 줄 = 한 케이스 ({len(ox_combo):,}가지)')
    run = 0
    body = []
    for key, cnt in ox_combo:
        run += cnt
        body.append([('O' if v else 'X') for v in key]
                    + [cnt, round(100.0 * cnt / n, 2), round(100.0 * run / n, 2)])
    ops.append(('tbl', [h for _, h in OX_FLAGS] + ['건수', '비율(%)', '누적(%)'],
                body, {'num': [len(ox_cols) + 1], 'ox': len(ox_cols)}))
    ops.append(('gap',))

    for tag, cols, label in [('6', COMBO_CORE, '핵심'), ('7', COMBO_FULL, '최대 세분화')]:
        combo = _combo(rows, cols)
        h2(f'§{tag}. {label} {len(cols)}축 조합 전수 ({len(combo):,}가지)')
        run, body = 0, []
        for key, cnt in combo:
            run += cnt
            body.append(list(key) + [cnt, round(100.0 * cnt / n, 2), round(100.0 * run / n, 2)])
        ops.append(('tbl', [dict(AXES)[c] for c in cols] + ['건수', '비율(%)', '누적(%)'],
                    body, {'num': [len(cols) + 1]}))
        ops.append(('gap',))

    pmap = {}
    for r in rows:
        pmap.setdefault((r['platform'], r['parent_key']), r)
    prows = list(pmap.values())
    h2(f'§8. 부모(포스트/영상) 단위 분포 — {len(prows):,}건')
    ops.append(('tbl', ['라벨 축', '값', '건수', '비율(%)'],
                [[dict(AXES)[col], k, v, round(100.0 * v / len(prows), 2)]
                 for col in ['platform', 'c_period', 'c_stage', 'c_multi']
                 for k, v in _dist(prows, col)], {'num': [3]}))
    ops.append(('gap',))

    bad = [r for r in rows if not r['ox_linkdone']]
    h2(f'§9. 미해결(done 아님) 진단 — {len(bad):,}건')
    ops.append(('tbl', ['미해결 세부사유', 'link_status', '공구기간', '플랫폼',
                        'link_location', '건수', '미해결중비율(%)'],
                [list(k) + [v, round(100.0 * v / len(bad), 2)]
                 for k, v in _combo(bad, ['c_unres', 'c_linkstatus', 'c_period',
                                          'platform', 'c_loc'])] if bad else [],
                {'num': [6]}))
    ops.append(('gap',))

    okl = [r for r in rows if r['ox_linkdone']]
    h2(f'§10. 링크확정(done) 건의 상세수집 진행 — {len(okl):,}건')
    ops.append(('tbl', ['도메인유형', 'URL_상세페이지여부', 'detail_status', '건수',
                        'done중비율(%)'],
                [list(k) + [v, round(100.0 * v / len(okl), 2)]
                 for k, v in _combo(okl, ['c_urlkind', 'c_depth', 'c_detail'])] if okl else [],
                {'num': [4]}))
    return ops


def write_xlsx_one(rows, path):
    """모든 표를 **시트 하나에** 세로로 쌓는다. 맨 위에 목차(§와 행 번호)를 두고,
    각 표는 굵은 머리글로 구분한다. 시트가 하나라 자동필터는 못 걸리는 대신
    Ctrl+F / 이름상자(예: A420)로 바로 점프해서 보게 만든 구성이다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    ops = _one_sheet_ops(rows)
    sections = [(t, i) for i, (kind, *rest) in enumerate(ops) if kind == 'h2'
                for t in [rest[0]]]

    # 1패스: 목차 블록 크기를 알아야 각 블록의 최종 행 번호가 정해진다.
    top = 2 + 1 + 1 + len(sections) + 2      # 제목·부제·목차머리글·목차본문·여백
    rowof, r = {}, top
    for idx, (kind, *rest) in enumerate(ops):
        rowof[idx] = r
        r += 1 if kind in ('h2', 'note', 'gap') else 1 + len(rest[1])

    wb = Workbook()
    ws = wb.active
    ws.title = '케이스_전수세분화'
    hdr_fill = PatternFill('solid', fgColor=_XL_HDR_FILL)
    sec_fill = PatternFill('solid', fgColor='FF5C7A8A')
    o_fill = PatternFill('solid', fgColor='FFD8EFD3')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.cell(row=1, column=1, value='공구왕 DB 케이스 전수 세분화 — 전체 표 (시트 1장)').font = \
        Font(bold=True, size=15)
    ws.cell(row=2, column=1,
            value=f"상품 행 {len(rows):,}건 · 라벨 축 {len(AXES)}개 · O/X 플래그 "
                  f"{len(OX_FLAGS)}개 · O = 값이 있음(NOT NULL·비어있지 않음)").font = \
        Font(italic=True, color='FF555555')
    for j, h in enumerate(['목차', '시작 행'], 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font, c.fill = Font(bold=True, color='FFFFFFFF'), hdr_fill
    for k, (t, idx) in enumerate(sections):
        ws.cell(row=5 + k, column=1, value=t)
        ws.cell(row=5 + k, column=2, value=rowof[idx] + 1)

    # 2패스: 실제 쓰기
    for idx, (kind, *rest) in enumerate(ops):
        r = rowof[idx]
        if kind == 'gap':
            continue
        if kind == 'h2':
            c = ws.cell(row=r, column=1, value=rest[0])
            c.font, c.fill = Font(bold=True, size=12, color='FFFFFFFF'), sec_fill
            continue
        if kind == 'note':
            ws.cell(row=r, column=1, value=rest[0]).font = Font(bold=True, color='FF2F4858')
            continue
        headers, body, opt = rest
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=j, value=h)
            c.font, c.fill, c.alignment = Font(bold=True, color='FFFFFFFF'), hdr_fill, ctr
        for i, line in enumerate(body, 1):
            for j, v in enumerate(line, 1):
                c = ws.cell(row=r + i, column=j, value=v)
                if j in opt.get('num', ()):
                    c.number_format = '#,##0'
                if opt.get('ox') and j <= opt['ox']:
                    c.alignment = Alignment(horizontal='center')
                    if v == 'O':
                        c.fill = o_fill
                if j == 1 and opt.get('boldfirstcol'):
                    c.font = Font(bold=True)
            if opt.get('boldlast') and i == len(body):
                for j in range(1, len(line) + 1):
                    ws.cell(row=r + i, column=j).font = Font(bold=True)

    # 열 폭 — 한 시트에 성격이 다른 표가 섞이므로 타협값(A는 라벨용으로 넓게)
    from openpyxl.utils import get_column_letter
    for i in range(1, max(ws.max_column, 24) + 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            40 if i == 1 else (20 if i <= 13 else 13)
    ws.freeze_panes = 'A5'
    ws.sheet_view.showGridLines = False

    pathlib.Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def emit_sql(path):
    """같은 라벨링 로직을 순수 SQL(뷰 + 예제 GROUP BY)로 뽑는다 — DB 클라이언트에서
    바로 돌려보고 싶을 때. 뷰 1개를 만들 뿐 데이터는 건드리지 않는다."""
    axis_cols = [c for c, _ in AXES]
    parts = [
        '-- gonggu/case_matrix.py --emit-sql 로 생성됨(직접 손으로 고치지 말 것).',
        '-- 읽기 전용 뷰 1개(v_gonggu_case_axes) + 예제 집계 쿼리.',
        '-- 뷰만 만들고 기존 테이블/데이터는 전혀 건드리지 않는다.',
        '',
        'CREATE OR REPLACE VIEW v_gonggu_case_axes AS',
        _axis_sql().strip() + ';',
        '',
        '-- ============ 예제 1) 축별 단독 분포 ============',
    ]
    for col, name in AXES:
        parts.append(f"-- {name}\nSELECT {col} AS `값`, COUNT(*) AS `건수` "
                     f"FROM v_gonggu_case_axes GROUP BY {col} ORDER BY `건수` DESC;\n")
    parts += [
        '-- ============ 예제 2) 2축 교차표 ============',
    ]
    for rax, cax, title in CROSSTABS:
        parts.append(f"-- {title}\nSELECT {rax}, {cax}, COUNT(*) AS `건수` "
                     f"FROM v_gonggu_case_axes GROUP BY {rax}, {cax} "
                     f"ORDER BY {rax}, `건수` DESC;\n")
    ox = [c for c, _ in OX_FLAGS]
    ox_sel = ',\n       '.join(f"IF({c}, 'O', 'X') AS `{h}`" for c, h in OX_FLAGS)
    parts += [
        '-- ============ 예제 3) O/X 조합 전수 (한 줄 = 한 케이스) ============',
        f"SELECT {ox_sel},\n       COUNT(*) AS `건수`\n  FROM v_gonggu_case_axes\n"
        f" GROUP BY {', '.join(ox)}\n ORDER BY `건수` DESC;",
        '',
        '-- ============ 예제 4) 최대 세분화 라벨 조합 전수 ============',
        f"SELECT {', '.join(axis_cols)}, COUNT(*) AS `건수`\n  FROM v_gonggu_case_axes\n"
        f" GROUP BY {', '.join(axis_cols)}\n ORDER BY `건수` DESC;",
        '',
        '-- ============ 예제 5) 미해결(done 아님)만 세부 사유별 ============',
        "SELECT c_unres, c_linkstatus, c_period, COUNT(*) AS `건수`\n"
        "  FROM v_gonggu_case_axes\n WHERE link_status IS NULL OR link_status <> 'done'\n"
        " GROUP BY c_unres, c_linkstatus, c_period\n ORDER BY `건수` DESC;",
        '',
    ]
    pathlib.Path(path).write_text('\n'.join(parts), encoding='utf-8')
    print(f'SQL 생성: {path}')


def main():
    ap = argparse.ArgumentParser(description='DB 케이스 전수 세분화 리포트(읽기 전용)')
    ap.add_argument('--limit', type=int, default=0, help='상품 행 N건만 (소량 확인용)')
    ap.add_argument('--emit-sql', metavar='PATH', nargs='?', const='queries/case_matrix.sql',
                    help='DB 접속 없이 같은 로직의 순수 SQL만 파일로 출력')
    ap.add_argument('--from-csv', metavar='PATH',
                    help='DB 대신 앞선 실행의 case_matrix_rows.csv에서 읽어 리포트만 다시 만든다')
    ap.add_argument('--out', metavar='DIR', help='출력 폴더(기본 data/output/case_matrix)')
    ap.add_argument('--no-xlsx', action='store_true', help='엑셀 워크북은 만들지 않는다')
    ap.add_argument('--one-sheet', action='store_true',
                    help='엑셀을 시트 20개가 아니라 시트 1장(목차+모든 표 세로 누적)으로')
    args = ap.parse_args()

    if args.emit_sql:
        emit_sql(ROOT / args.emit_sql if not pathlib.Path(args.emit_sql).is_absolute()
                 else args.emit_sql)
        return

    out_dir = pathlib.Path(args.out) if args.out else OUT_DIR

    if args.from_csv:
        rows = load_rows_from_csv(args.from_csv)
        print(f'CSV에서 {len(rows):,}행 복원: {args.from_csv}')
    else:
        sql = _axis_sql()
        if args.limit:
            sql += f'\n LIMIT {int(args.limit)}'
        conn = connect_dst()
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                cur.execute(sql)
                rows = cur.fetchall()
        finally:
            conn.close()

    if not rows:
        print('대상 행이 없습니다.', file=sys.stderr)
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    md = build_report(rows)
    (out_dir / 'case_matrix.md').write_text(md, encoding='utf-8')
    if not args.from_csv:          # rows.csv를 입력으로 썼으면 그 파일을 덮어쓰지 않는다
        write_csvs(rows, out_dir)
    xlsx = None
    if not args.no_xlsx:
        if args.one_sheet:
            xlsx = write_xlsx_one(rows, out_dir / 'case_matrix_one_sheet.xlsx')
        else:
            xlsx = write_xlsx(rows, out_dir / 'case_matrix.xlsx')

    print(f'상품 행 {len(rows):,}건 분석 완료')
    for col, name in AXES:
        d = _dist(rows, col)
        head = ', '.join(f'{k}={v:,}' for k, v in d[:5])
        print(f'  - {name}: {len(d)}종 — {head}{" ..." if len(d) > 5 else ""}')
    print(f'  - O/X 조합 실존: {len(_combo(rows, [c for c, _ in OX_FLAGS])):,}가지')
    print(f'  - 최대세분화 라벨 조합 실존: {len(_combo(rows, COMBO_FULL)):,}가지')
    print(f'출력: {out_dir}/case_matrix.md'
          + ('' if args.from_csv else ', case_matrix_{rows,ox,combo,dist}.csv')
          + (f', {xlsx.name}' if xlsx else ''))


if __name__ == '__main__':
    main()
