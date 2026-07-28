#!/usr/bin/env python3
"""classify_category.py 결과(JSONL)에서 category=="미분류"인 것만 뽑아 CSV+엑셀로 저장한다.
DB에서 원본 포스트/영상 URL(그리고 있으면 제품 구매링크)도 같이 붙여서, 그 URL만 클릭하면
바로 어떤 제품인지 확인할 수 있게 한다. 엑셀 쪽은 url/purchase_url을 클릭 가능한 링크로 만든다.

사용법:
    python3 scripts/export_unclassified.py                      # 기본 입력/출력 경로
    python3 scripts/export_unclassified.py <결과.jsonl> <출력.csv>
결과: <출력.csv>와 같은 이름의 .xlsx(예: gonggu_category_unclassified.xlsx)도 같이 저장.
"""
import csv
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Font

from common import connect_dst

IN_DEFAULT = pathlib.Path.home() / 'Desktop' / 'gonggu_category_result.jsonl'
OUT_DEFAULT = pathlib.Path.home() / 'Desktop' / 'gonggu_category_unclassified.csv'
HEADERS = ['product_name', 'title', 'description', 'confidence', 'reason', 'url', 'purchase_url']


def _load_jsonl(path):
    with open(path, encoding='utf-8') as f:
        return [json.loads(line) for line in f if line.strip()]


def _chunks(seq, n=1000):
    seq = list(seq)
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def fetch_post_urls(conn, post_ids):
    out = {}
    with conn.cursor() as cur:
        for chunk in _chunks(set(post_ids)):
            fmt = ','.join(['%s'] * len(chunk))
            cur.execute(f'SELECT post_id, url FROM gonggu_post WHERE post_id IN ({fmt})', chunk)
            for r in cur.fetchall():
                out[r['post_id']] = r['url'] or ''
    return out


def fetch_video_urls(conn, video_ids):
    out = {}
    with conn.cursor() as cur:
        for chunk in _chunks(set(video_ids)):
            fmt = ','.join(['%s'] * len(chunk))
            cur.execute(f'SELECT video_id, video_url FROM gonggu_video WHERE video_id IN ({fmt})', chunk)
            for r in cur.fetchall():
                out[r['video_id']] = r['video_url'] or ''
    return out


def fetch_purchase_urls(conn, table, id_col, pairs):
    """pairs: [(content_id, product_name), ...] -> {(content_id, product_name): candidate_url}"""
    out = {}
    with conn.cursor() as cur:
        for chunk in _chunks(pairs):
            conds = ' OR '.join([f'({id_col} = %s AND product_name = %s)'] * len(chunk))
            params = [v for pair in chunk for v in pair]
            cur.execute(f'SELECT {id_col}, product_name, candidate_url FROM {table} WHERE {conds}', params)
            for r in cur.fetchall():
                out[(r[id_col], r['product_name'])] = r['candidate_url'] or ''
    return out


def main():
    in_path = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else IN_DEFAULT
    out_path = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else OUT_DEFAULT

    rows = [r for r in _load_jsonl(in_path) if r.get('category') == '미분류']
    print(f'미분류 {len(rows)}건 (전체 {in_path.name} 기준)')

    ig_rows = [r for r in rows if r['platform'] == '인스타']
    yt_rows = [r for r in rows if r['platform'] == '유튜브']

    conn = connect_dst()
    try:
        post_urls = fetch_post_urls(conn, [r['content_id'] for r in ig_rows])
        video_urls = fetch_video_urls(conn, [r['content_id'] for r in yt_rows])
        ig_purchase = fetch_purchase_urls(
            conn, 'gonggu_post_product', 'post_id',
            [(r['content_id'], r['product_name']) for r in ig_rows])
        yt_purchase = fetch_purchase_urls(
            conn, 'gonggu_video_product', 'video_id',
            [(r['content_id'], r['product_name']) for r in yt_rows])
    finally:
        conn.close()

    table = []
    for r in rows:
        if r['platform'] == '인스타':
            url = post_urls.get(r['content_id'], '')
            purchase_url = ig_purchase.get((r['content_id'], r['product_name']), '')
        else:
            url = video_urls.get(r['content_id'], '')
            purchase_url = yt_purchase.get((r['content_id'], r['product_name']), '')
        table.append([
            r.get('product_name') or '', r.get('title') or '', r.get('description') or '',
            r.get('confidence'), r.get('reason') or '', url, purchase_url,
        ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(table)
    print(f'CSV 저장 완료 -> {out_path}')

    xlsx_path = out_path.with_suffix('.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = '미분류'
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    url_col, purchase_col = HEADERS.index('url') + 1, HEADERS.index('purchase_url') + 1
    for row in table:
        ws.append(row)
        r = ws.max_row
        for col in (url_col, purchase_col):
            val = ws.cell(row=r, column=col).value
            if val:
                ws.cell(row=r, column=col).hyperlink = val
                ws.cell(row=r, column=col).font = Font(color='0563C1', underline='single')
    widths = [30, 25, 50, 10, 35, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f'엑셀 저장 완료 -> {xlsx_path}')


if __name__ == '__main__':
    main()
